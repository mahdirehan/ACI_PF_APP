"""
One-off migration script: rename PI (Priority Index) -> PF (Priority Factor)
in every Excel workbook in the project.

Replacements (applied to every cell value, every sheet, every workbook):
  * Whole-cell value "PI"            -> "PF"
  * Whole-cell value "PI_Computed"   -> "PF_Computed"
  * Substring     "Priority Index"   -> "Priority Factor"
  * Substring     "(PI)"             -> "(PF)"
  * Whole-cell value "pi"            -> "pf"

The script also renames matching sheet/tab names using the same rules.

Original workbooks are backed up alongside the original as
  <name>.bak_pi2pf.xlsx
before any modification.

Run from the project root or from anywhere - paths are absolute below.
"""

from __future__ import annotations

import shutil
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
RTA_ROOT = PROJECT_ROOT / "rta-asset-management"

TARGET_FILES = [
    PROJECT_ROOT / "1. Calculations for Priority Index.xlsx",
    PROJECT_ROOT / "MASTER_EXCEL_E11.xlsx",
    PROJECT_ROOT / "Master_Priority_Output.xlsx",
    PROJECT_ROOT / "RoW Assets_ACI Sheet_20-05-2025.xlsx",
    RTA_ROOT / "test_report_2026-05-18_165658.xlsx",
]

WHOLE_CELL_REPLACEMENTS = {
    "PI": "PF",
    "PI_Computed": "PF_Computed",
    "pi": "pf",
    "PI Tests": "PF Tests",
}

SUBSTRING_REPLACEMENTS = [
    ("Priority Index", "Priority Factor"),
    ("(PI)", "(PF)"),
    ("PI =", "PF ="),
    ("PI Calculation", "PF Calculation"),
    ("PIResult", "PFResult"),
]


def _rewrite_value(value: Any) -> Any:
    """Apply renaming rules to a single cell value."""
    if not isinstance(value, str):
        return value

    stripped = value.strip()
    if stripped in WHOLE_CELL_REPLACEMENTS:
        return WHOLE_CELL_REPLACEMENTS[stripped]

    new_value = value
    for old, new in SUBSTRING_REPLACEMENTS:
        if old in new_value:
            new_value = new_value.replace(old, new)

    return new_value


def _rewrite_sheet_name(name: str) -> str:
    """Apply renaming rules to a sheet name."""
    new_name = name
    if new_name.strip() in WHOLE_CELL_REPLACEMENTS:
        new_name = WHOLE_CELL_REPLACEMENTS[new_name.strip()]
    for old, new in SUBSTRING_REPLACEMENTS:
        if old in new_name:
            new_name = new_name.replace(old, new)
    return new_name


def process_workbook(path: Path) -> dict[str, int]:
    """
    Apply PI -> PF replacements throughout one workbook in place.

    Returns a dict of statistics: cells_changed, sheets_renamed.
    """
    stats = {"cells_changed": 0, "sheets_renamed": 0}

    if not path.exists():
        print(f"  ! Skipping (not found): {path}")
        return stats

    backup_path = path.with_suffix(".bak_pi2pf.xlsx")
    if not backup_path.exists():
        shutil.copy2(path, backup_path)
        print(f"  - Backup created: {backup_path.name}")

    print(f"  Loading: {path.name}")
    try:
        wb = load_workbook(path)
    except Exception as exc:
        print(f"  ! Failed to load {path.name}: {exc}")
        return stats

    for ws in wb.worksheets:
        original_title = ws.title
        new_title = _rewrite_sheet_name(original_title)
        if new_title != original_title:
            try:
                ws.title = new_title
                stats["sheets_renamed"] += 1
                print(f"    Sheet renamed: '{original_title}' -> '{new_title}'")
            except Exception as exc:
                print(f"    ! Could not rename sheet '{original_title}': {exc}")

        for row in ws.iter_rows():
            for cell in row:
                old = cell.value
                if old is None:
                    continue
                new = _rewrite_value(old)
                if new != old:
                    cell.value = new
                    stats["cells_changed"] += 1

    try:
        wb.save(path)
        print(f"  Saved: {path.name}")
    except Exception as exc:
        print(f"  ! Failed to save {path.name}: {exc}")

    return stats


def main() -> int:
    print("=" * 70)
    print("PI -> PF Excel migration")
    print("=" * 70)

    total_cells = 0
    total_sheets = 0
    for path in TARGET_FILES:
        print(f"\n[{path.name}]")
        stats = process_workbook(path)
        total_cells += stats["cells_changed"]
        total_sheets += stats["sheets_renamed"]

    print("\n" + "=" * 70)
    print(f"Total cells changed:  {total_cells}")
    print(f"Total sheets renamed: {total_sheets}")
    print("Backups saved alongside originals as '<name>.bak_pi2pf.xlsx'.")
    print("=" * 70)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
