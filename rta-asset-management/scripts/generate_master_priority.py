"""
Master Priority Output Generator.

Reads asset inventory data from multiple Excel files (with each asset type
typically living in its own sheet/file), consolidates every row into a single
unified schema, computes any missing Priority Factor (PF) values, and writes a
professionally-formatted master Excel file ranked from highest to lowest PF.

Output: Master_Priority_Output.xlsx with three sheets:
  1. Priority Ranking  - the main deliverable (ranked, formatted, filtered)
  2. Summary Statistics - aggregate stats via live Excel formulas
  3. Source Mapping     - which file/sheet contributed which asset type

Usage:
    python -m scripts.generate_master_priority
    python scripts/generate_master_priority.py
    python scripts/generate_master_priority.py --output ./my_output.xlsx

The script auto-discovers all .xlsx files in the parent directory by default.
"""

from __future__ import annotations

import argparse
import re
import sys
import time
from collections import Counter
from dataclasses import dataclass, field
from datetime import datetime, date
from pathlib import Path
from typing import Any, Optional

import xlsxwriter
from openpyxl import load_workbook


SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_DIR = SCRIPT_DIR.parent
DEFAULT_INPUT_DIR = PROJECT_DIR.parent
DEFAULT_OUTPUT_FILE = DEFAULT_INPUT_DIR / "Master_Priority_Output.xlsx"

if str(PROJECT_DIR) not in sys.path:
    sys.path.insert(0, str(PROJECT_DIR))

try:
    from domain.pf.cf import calculate_cf as _domain_calculate_cf
    from domain.pf.rf import get_rf_for_asset as _domain_get_rf
    from domain.pf.flf import calculate_flf as _domain_calculate_flf
    _DOMAIN_AVAILABLE = True
except Exception:
    _DOMAIN_AVAILABLE = False

DEFAULT_PF_WEIGHTS = {"cf": 0.60, "rf": 0.20, "flf": 0.20}

HEADER_SEARCH_ROWS = 12
MIN_DATA_FIELDS_FOR_HEADER = 3


def _log(message: str) -> None:
    """Print with immediate flush so progress is visible during long sheets."""
    print(message, flush=True)

NON_INVENTORY_SHEET_HINTS = {
    "combined data",
    "summary",
    "sheet2",
    "sheet3",
    "aci_calculation sheet",
    "priority index cal sheet",
    "priority factor cal sheet",
    "condition factor",
    "risk factor",
    "functional life factor",
    "junctions_tabletoexcel",
}

CANONICAL_FIELDS = [
    "Asset_ID",
    "Asset_Type",
    "Road_Number",
    "Road_Name_EN",
    "Road_Name_AR",
    "Road_Classification",
    "Road_Surface_Type",
    "Route_ID",
    "Route_Name",
    "Direction",
    "Lane_Route_ID",
    "Chainage_Start",
    "Chainage_End",
    "Offset",
    "Roadside_Location",
    "Description",
    "Inspection_Date",
    "ACI",
    "CF",
    "FLF",
    "RF",
    "PF",
    "ACI_Rating",
    "PF_Computed",
    "Source_Sheet",
    "Source_File",
]

COLUMN_ALIASES: dict[str, str] = {
    "objectid": "Asset_ID",
    "objectid *": "Asset_ID",
    "object_id": "Asset_ID",
    "asset_id": "Asset_ID",
    "element_id": "Asset_ID",
    "unique_asset_name": "Asset_ID",
    "asset_name": "Asset_Type",
    "asset_type": "Asset_Type",
    "feature_name": "Asset_Type",
    "type": "Asset_Type",
    "road_number": "Road_Number",
    "roadname_en": "Road_Name_EN",
    "road_name_en": "Road_Name_EN",
    "roadname_ar": "Road_Name_AR",
    "road_name_ar": "Road_Name_AR",
    "road_classification": "Road_Classification",
    "road_surface_type": "Road_Surface_Type",
    "route_id": "Route_ID",
    "route_name": "Route_Name",
    "direction": "Direction",
    "lane_route_id": "Lane_Route_ID",
    "lrs_chainage": "Chainage_Start",
    "lrs_chainage_start": "Chainage_Start",
    "chainage": "Chainage_Start",
    "chainage_start": "Chainage_Start",
    "lrs_chainage_end": "Chainage_End",
    "chainage_end": "Chainage_End",
    "offset": "Offset",
    "roadsidelocation": "Roadside_Location",
    "roadside_location": "Roadside_Location",
    "road_side_location": "Roadside_Location",
    "survey_date": "Inspection_Date",
    "surveydate": "Inspection_Date",
    "inspection_date": "Inspection_Date",
    "aci": "ACI",
    "aci_1": "ACI",
    "cf": "CF",
    "rf": "RF",
    "flf": "FLF",
    "pi": "PF",
    "pf": "PF",
    "aci_rating": "ACI_Rating",
}

DEFECT_KEYWORD = "DEFECTS"


@dataclass
class AssetRow:
    """A single normalized asset row from any source sheet."""

    canonical: dict[str, Any] = field(default_factory=dict)
    extras: dict[str, Any] = field(default_factory=dict)


@dataclass
class SheetExtract:
    """Container for everything extracted from one sheet."""

    file_name: str
    sheet_name: str
    rows: list[AssetRow]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _normalize_header(raw: Any) -> str:
    """Normalize a header cell into a lookup key."""
    if raw is None:
        return ""
    text = str(raw).strip()
    text = re.sub(r"\s+", " ", text)
    text = text.replace("\n", " ").replace("\r", " ")
    text = text.strip(" *").strip()
    key = text.lower()
    key = re.sub(r"[^a-z0-9]+", "_", key).strip("_")
    return key


def _looks_like_header_row(row: tuple[Any, ...]) -> bool:
    """Heuristic: a header row has several non-empty short text cells."""
    text_cells = [c for c in row if isinstance(c, str) and 1 <= len(c.strip()) <= 80]
    non_text_cells = [c for c in row if c is not None and not isinstance(c, str)]
    return len(text_cells) >= MIN_DATA_FIELDS_FOR_HEADER and len(text_cells) >= len(non_text_cells)


def _find_header_row(ws: Any) -> Optional[int]:
    """
    Find the row index (1-based) of the most likely header.

    Strategy: scan first ``HEADER_SEARCH_ROWS`` rows; prefer the one containing
    an ACI/PF/CF/RF/FLF column. Fall back to the first row that looks textual.
    """
    best: Optional[tuple[int, int]] = None
    fallback: Optional[int] = None
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=1, max_row=HEADER_SEARCH_ROWS, values_only=True),
        start=1,
    ):
        if not _looks_like_header_row(row):
            continue
        if fallback is None:
            fallback = row_idx
        hit_score = 0
        for cell in row:
            if cell is None:
                continue
            key = _normalize_header(cell)
            if key in {"aci", "pi", "pf", "cf", "rf", "flf", "aci_rating"}:
                hit_score += 2
            elif "aci" in key or "pi" in key or "pf" in key or "objectid" in key:
                hit_score += 1
        if hit_score > 0 and (best is None or hit_score > best[1]):
            best = (row_idx, hit_score)
    if best is not None:
        return best[0]
    return fallback


def _is_meaningful(value: Any) -> bool:
    """Return True if a cell value is non-empty/meaningful."""
    if value is None:
        return False
    if isinstance(value, str):
        return value.strip() != ""
    return True


def _to_float(value: Any) -> Optional[float]:
    """Coerce a value to float, tolerating string-encoded numbers."""
    if value is None:
        return None
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        text = value.strip().replace(",", "")
        if not text or text in {"-", "N/A", "NA", "None"}:
            return None
        try:
            return float(text)
        except ValueError:
            return None
    return None


def _to_str(value: Any) -> str:
    """Coerce to a clean string, handling dates and numbers neatly."""
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return f"{value:.4f}".rstrip("0").rstrip(".")
    return str(value).strip()


def _classify_column(header_key: str) -> tuple[str, Optional[str]]:
    """
    Classify a column header into (kind, canonical_name).

    ``kind`` is one of: ``canonical``, ``defect``, ``rating``, ``extra``.
    """
    if not header_key:
        return ("extra", None)
    if header_key in COLUMN_ALIASES:
        return ("canonical", COLUMN_ALIASES[header_key])
    if header_key.endswith("_defects") or header_key == "defects":
        return ("defect", header_key)
    if header_key.endswith("_rating") and header_key != "aci_rating":
        return ("rating", header_key)
    return ("extra", header_key)


def _humanize(key: str) -> str:
    """Turn ``foundation_defects`` into ``Foundation Defects``."""
    return key.replace("_", " ").strip().title()


# ---------------------------------------------------------------------------
# Sheet extraction
# ---------------------------------------------------------------------------


def is_inventory_sheet(sheet_name: str, headers: list[str]) -> bool:
    """A sheet is treated as inventory if it has an ACI-like column and rows."""
    lowered = sheet_name.strip().lower()
    if lowered in NON_INVENTORY_SHEET_HINTS:
        return False
    keys = {_normalize_header(h) for h in headers}
    return any(k in {"aci", "aci_1"} for k in keys) or any(k.startswith("aci") for k in keys)


def extract_sheet(
    ws: Any,
    file_name: str,
    sheet_name: str,
) -> Optional[SheetExtract]:
    """Extract normalized rows from a single worksheet."""
    header_row = _find_header_row(ws)
    if header_row is None:
        return None

    raw_headers: list[Any] = []
    for row in ws.iter_rows(
        min_row=header_row, max_row=header_row, values_only=True
    ):
        raw_headers = list(row)
        break
    if not raw_headers:
        return None

    headers_text = [str(h) if h is not None else "" for h in raw_headers]
    if not is_inventory_sheet(sheet_name, headers_text):
        return None

    column_map: list[tuple[int, str, Optional[str], str]] = []
    for col_idx, raw in enumerate(raw_headers):
        if raw is None or str(raw).strip() == "":
            continue
        key = _normalize_header(raw)
        kind, canonical_or_key = _classify_column(key)
        column_map.append((col_idx, key, canonical_or_key, kind))

    rows_out: list[AssetRow] = []
    for r in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not any(_is_meaningful(c) for c in r):
            continue
        canonical: dict[str, Any] = {}
        defects: list[str] = []
        extras: dict[str, Any] = {}
        for col_idx, key, target, kind in column_map:
            if col_idx >= len(r):
                continue
            value = r[col_idx]
            if not _is_meaningful(value):
                continue
            if kind == "canonical":
                canonical[target] = value  # type: ignore[index]
            elif kind == "defect":
                text = _to_str(value)
                if text and text.lower() not in {"no defect found", "none", "-"}:
                    defects.append(f"{_humanize(key)}: {text}")
                extras[_humanize(key)] = value
            elif kind == "rating":
                extras[_humanize(key)] = value
            else:
                extras[_humanize(key)] = value

        if defects:
            canonical.setdefault("Description", "; ".join(defects))
        if "Asset_Type" not in canonical:
            canonical["Asset_Type"] = _infer_asset_type_from_sheet(sheet_name)
        canonical["Source_Sheet"] = sheet_name
        canonical["Source_File"] = file_name

        if not _has_any_useful_value(canonical):
            continue

        rows_out.append(AssetRow(canonical=canonical, extras=extras))

    if not rows_out:
        return None
    return SheetExtract(file_name=file_name, sheet_name=sheet_name, rows=rows_out)


def _infer_asset_type_from_sheet(sheet_name: str) -> str:
    """Pretty-print an asset type from the sheet name."""
    cleaned = re.sub(r"(?i)_tabletoexcel", "", sheet_name)
    cleaned = re.sub(r"(?i)_tabletoexce$", "", cleaned)
    cleaned = cleaned.replace("_", " ").strip()
    return cleaned.title() if cleaned else sheet_name


def _has_any_useful_value(canonical: dict[str, Any]) -> bool:
    """Row is useful if it carries an ID, ACI, or a type-and-defect."""
    if _is_meaningful(canonical.get("Asset_ID")):
        return True
    if _to_float(canonical.get("ACI")) is not None:
        return True
    if _is_meaningful(canonical.get("Description")):
        return True
    return False


# ---------------------------------------------------------------------------
# PF calculation
# ---------------------------------------------------------------------------


def _normalize_asset_type_key(asset_type: Any) -> str:
    """Map a free-form asset type label to the lookup-table key format."""
    text = _to_str(asset_type).upper().strip()
    text = re.sub(r"[^A-Z0-9]+", "_", text).strip("_")
    return text


def _derive_factors_from_aci(
    aci: float, asset_type: Any
) -> tuple[Optional[float], Optional[float], Optional[float]]:
    """
    Use the domain CF/RF/FLF formulas to fill in factors from ACI + asset type.

    Returns (cf, rf, flf). Any may be None if the lookup fails. CF is always
    computable from ACI alone.
    """
    cf: Optional[float] = None
    rf: Optional[float] = None
    flf: Optional[float] = None
    if _DOMAIN_AVAILABLE:
        try:
            cf = _domain_calculate_cf(aci)
        except Exception:
            cf = None
        key = _normalize_asset_type_key(asset_type)
        if key:
            try:
                rf = _domain_get_rf(key)
            except Exception:
                rf = None
            try:
                flf = _domain_calculate_flf(key, aci)
                if flf == 0.0 and key.upper() not in {""}:
                    pass
            except Exception:
                flf = None
    else:
        if aci >= 80:
            cf = 0.0
            flf = 0.0
        else:
            cf = (80 - aci) / 80 * 100
    return cf, rf, flf


def compute_missing_pf(
    rows: list[AssetRow], weights: dict[str, float] = DEFAULT_PF_WEIGHTS
) -> int:
    """
    Fill in missing PF values.

    Strategy (in order):
      1. If PF is already present, keep it.
      2. If CF, RF, FLF are all present, apply the weighted formula directly.
      3. Otherwise derive missing factors from ACI + Asset_Type via the
         domain modules (CF from ACI; RF from rf_scores; FLF from curve).
         Any factor that still cannot be derived is treated as 0 so the
         caller can rank approximately rather than dropping the row.

    Returns the count of rows for which PF was computed.

    The formula matches ``domain/pf/aggregator.py``:
        PF = w_cf * CF + w_rf * RF + w_flf * FLF
    """
    w_cf = weights["cf"]
    w_rf = weights["rf"]
    w_flf = weights["flf"]
    computed = 0
    for row in rows:
        pf_existing = _to_float(row.canonical.get("PF"))
        if pf_existing is not None:
            row.canonical["PF"] = round(pf_existing, 4)
            row.canonical["PF_Computed"] = False
            continue
        cf = _to_float(row.canonical.get("CF"))
        rf = _to_float(row.canonical.get("RF"))
        flf = _to_float(row.canonical.get("FLF"))
        aci = _to_float(row.canonical.get("ACI"))
        asset_type = row.canonical.get("Asset_Type")

        if (cf is None or rf is None or flf is None) and aci is not None:
            derived_cf, derived_rf, derived_flf = _derive_factors_from_aci(
                aci, asset_type
            )
            if cf is None and derived_cf is not None:
                cf = derived_cf
                row.canonical["CF"] = round(derived_cf, 4)
            if rf is None and derived_rf is not None:
                rf = derived_rf
                row.canonical["RF"] = round(derived_rf, 4)
            if flf is None and derived_flf is not None:
                flf = derived_flf
                row.canonical["FLF"] = round(derived_flf, 4)

        if cf is not None and rf is not None and flf is not None:
            pf = w_cf * cf + w_rf * rf + w_flf * flf
            row.canonical["PF"] = round(pf, 4)
            row.canonical["PF_Computed"] = True
            computed += 1
        else:
            row.canonical["PF_Computed"] = False
    return computed


# ---------------------------------------------------------------------------
# Consolidation
# ---------------------------------------------------------------------------


def consolidate(extracts: list[SheetExtract]) -> tuple[list[str], list[AssetRow]]:
    """
    Merge all extracts into a single sorted list of rows.

    Returns (extra_columns_in_order, all_rows). ``extra_columns_in_order`` is
    the union of extra column names sorted by first appearance and frequency
    so the output is deterministic.
    """
    all_rows: list[AssetRow] = []
    extra_first_seen: dict[str, int] = {}
    extra_counts: Counter[str] = Counter()
    seq = 0
    for extract in extracts:
        for row in extract.rows:
            all_rows.append(row)
            for key in row.extras:
                if key not in extra_first_seen:
                    extra_first_seen[key] = seq
                    seq += 1
                extra_counts[key] += 1

    extra_columns = sorted(
        extra_first_seen.keys(),
        key=lambda k: (-extra_counts[k], extra_first_seen[k]),
    )

    compute_missing_pf(all_rows)

    def sort_key(row: AssetRow) -> tuple[float, float, str]:
        pf = _to_float(row.canonical.get("PF"))
        aci = _to_float(row.canonical.get("ACI"))
        # None PF sorts last (treat as -inf), None ACI sorts last (treat as +inf)
        pf_k = pf if pf is not None else float("-inf")
        aci_k = aci if aci is not None else float("inf")
        return (-pf_k, aci_k, _to_str(row.canonical.get("Asset_Type")))

    all_rows.sort(key=sort_key)
    return extra_columns, all_rows


# ---------------------------------------------------------------------------
# Output formatting (xlsxwriter)
# ---------------------------------------------------------------------------
#
# We use xlsxwriter (not openpyxl) for output because:
#   * It is dramatically faster on large datasets (~40k+ rows).
#   * It produces uncorrupted zip archives where openpyxl's non-streaming
#     writer truncates.
#   * Format objects are interned by the workbook, so cell-level styling is
#     cheap (each cell stores a reference, not a copy).


HEADER_BG = "#1E3C72"
HEADER_FG = "#FFFFFF"
ALT_BG = "#F5F5F5"
TITLE_COLOR = "#1E3C72"
SUBTITLE_COLOR = "#2A5298"
SECTION_COLOR = "#1E3C72"
BORDER_COLOR = "#BFBFBF"

PRIORITY_COLUMNS = [
    "Rank",
    "Asset_ID",
    "Asset_Type",
    "Road_Number",
    "Road_Name_EN",
    "Road_Name_AR",
    "Road_Classification",
    "Route_Name",
    "Direction",
    "Chainage_Start",
    "Chainage_End",
    "Offset",
    "Roadside_Location",
    "Description",
    "Inspection_Date",
    "ACI",
    "CF",
    "FLF",
    "RF",
    "PF",
    "ACI_Rating",
    "PF_Computed",
    "Source_Sheet",
    "Source_File",
]

FACTOR_COLS = {"ACI", "CF", "FLF", "RF", "PF"}


def _coerce_for_excel(value: Any) -> Any:
    """Make a value safe for xlsxwriter: numbers as floats, others as strings."""
    if value is None:
        return None
    if isinstance(value, (int, float, bool, datetime, date)):
        return value
    if isinstance(value, str):
        return value
    return str(value)


def _deduplicate_columns(columns: list[str]) -> list[str]:
    """Ensure column names are unique by suffixing duplicates."""
    seen: dict[str, int] = {}
    out: list[str] = []
    for name in columns:
        if name not in seen:
            seen[name] = 1
            out.append(name)
        else:
            seen[name] += 1
            out.append(f"{name} ({seen[name]})")
    return out


def _escape_for_formula(text: str) -> str:
    """Escape double quotes for embedding in an Excel string literal."""
    return text.replace('"', '""')


def _xl_col_letter(idx_zero_based: int) -> str:
    """Convert a 0-based column index to its Excel column letter (A, B, ..., AA)."""
    n = idx_zero_based
    letters = ""
    while True:
        n, rem = divmod(n, 26)
        letters = chr(ord("A") + rem) + letters
        if n == 0:
            break
        n -= 1
    return letters


def _build_formats(wb: Any) -> dict[str, Any]:
    """Pre-create every xlsxwriter Format object the output needs."""
    base = {"font_name": "Arial", "font_size": 10, "border": 1, "border_color": BORDER_COLOR}
    header = {
        **base,
        "bold": True,
        "font_color": HEADER_FG,
        "bg_color": HEADER_BG,
        "align": "center",
        "valign": "vcenter",
        "text_wrap": True,
    }
    return {
        "title": wb.add_format(
            {"font_name": "Arial", "font_size": 14, "bold": True, "font_color": TITLE_COLOR}
        ),
        "subtitle": wb.add_format(
            {"font_name": "Arial", "font_size": 12, "bold": True, "font_color": SUBTITLE_COLOR}
        ),
        "section": wb.add_format(
            {"font_name": "Arial", "font_size": 11, "bold": True, "font_color": SECTION_COLOR}
        ),
        "italic_grey": wb.add_format(
            {"font_name": "Arial", "font_size": 10, "italic": True, "font_color": "#666666"}
        ),
        "bold_small": wb.add_format(
            {"font_name": "Arial", "font_size": 10, "bold": True}
        ),
        "header": wb.add_format(header),
        "body": wb.add_format(base),
        "body_alt": wb.add_format({**base, "bg_color": ALT_BG}),
        "body_center": wb.add_format({**base, "align": "center"}),
        "body_center_alt": wb.add_format({**base, "align": "center", "bg_color": ALT_BG}),
        "body_num": wb.add_format({**base, "align": "right", "num_format": "0.00"}),
        "body_num_alt": wb.add_format(
            {**base, "align": "right", "num_format": "0.00", "bg_color": ALT_BG}
        ),
        "body_pct": wb.add_format({**base, "num_format": "0.0%"}),
        "summary_label": wb.add_format(
            {"font_name": "Arial", "font_size": 10, "bold": True, "border": 1, "border_color": BORDER_COLOR}
        ),
    }


def _write_value(ws: Any, row: int, col: int, value: Any, fmt: Any) -> int:
    """
    Write a value through xlsxwriter, choosing the right call for each type.

    Returns the on-screen length (used to update auto-fit width tracking).
    """
    if value is None:
        ws.write_blank(row, col, None, fmt)
        return 0
    if isinstance(value, str):
        if value.startswith("="):
            ws.write_formula(row, col, value, fmt)
            return 0
        ws.write_string(row, col, value, fmt)
        return len(value)
    if isinstance(value, bool):
        ws.write_boolean(row, col, value, fmt)
        return 5
    if isinstance(value, (int, float)):
        ws.write_number(row, col, value, fmt)
        return len(_to_str(value))
    if isinstance(value, (datetime, date)):
        ws.write_datetime(row, col, value, fmt)
        return 10
    text = _to_str(value)
    ws.write_string(row, col, text, fmt)
    return len(text)


def write_priority_sheet(
    ws: Any,
    fmts: dict[str, Any],
    full_columns: list[str],
    rows: list[AssetRow],
) -> None:
    """
    Write Sheet 1: Priority Ranking with professional formatting (xlsxwriter).

    Layout:
      * Header row 1 with dark fill + white bold text, frozen at top.
      * Data rows with alternating shading and right-aligned numeric cells.
      * Filter dropdowns on every column (auto-filter range).
      * 3-color conditional scales on the PF and ACI columns.
      * Auto-fitted column widths capped at 40 characters.

    The Rank column is populated with the ``=ROW()-1`` formula so it stays
    correct even after the user filters or sorts the table.
    """
    n_rows = len(rows)
    n_cols = len(full_columns)
    cap = 40

    max_widths: list[int] = [len(str(name)) for name in full_columns]

    for col_idx, name in enumerate(full_columns):
        ws.write_string(0, col_idx, name, fmts["header"])

    factor_set = {i for i, name in enumerate(full_columns) if name in FACTOR_COLS}
    rank_index = full_columns.index("Rank") if "Rank" in full_columns else -1
    center_set = {
        i
        for i, name in enumerate(full_columns)
        if name in {"Rank", "PF_Computed"}
    }

    canonical_pairs = [
        (i, name)
        for i, name in enumerate(full_columns)
        if (name in CANONICAL_FIELDS or name == "PF_Computed") and i != rank_index
    ]
    extra_pairs = [
        (i, name)
        for i, name in enumerate(full_columns)
        if i != rank_index
        and not (name in CANONICAL_FIELDS or name == "PF_Computed")
    ]

    body = fmts["body"]
    body_alt = fmts["body_alt"]
    body_center = fmts["body_center"]
    body_center_alt = fmts["body_center_alt"]
    body_num = fmts["body_num"]
    body_num_alt = fmts["body_num_alt"]

    for r_idx, asset in enumerate(rows, start=1):
        is_alt = r_idx % 2 == 0
        plain_fmt = body_alt if is_alt else body
        center_fmt = body_center_alt if is_alt else body_center
        num_fmt = body_num_alt if is_alt else body_num

        row_values: list[Any] = [None] * n_cols
        if rank_index >= 0:
            row_values[rank_index] = "=ROW()-1"
        for i, name in canonical_pairs:
            raw = asset.canonical.get(name)
            if i in factor_set:
                raw = _to_float(raw)
            row_values[i] = _coerce_for_excel(raw)
        for i, name in extra_pairs:
            row_values[i] = _coerce_for_excel(asset.extras.get(name))

        for i, value in enumerate(row_values):
            if i in factor_set:
                fmt = num_fmt
            elif i in center_set:
                fmt = center_fmt
            else:
                fmt = plain_fmt
            length = _write_value(ws, r_idx, i, value, fmt)
            if length > max_widths[i]:
                max_widths[i] = min(length, cap)

    for col_idx, width in enumerate(max_widths):
        ws.set_column(col_idx, col_idx, min(max(width + 2, 8), cap))

    ws.freeze_panes(1, 1)

    last_row = max(1, n_rows)
    ws.autofilter(0, 0, last_row, n_cols - 1)

    if n_rows > 0 and "PF" in full_columns and "ACI" in full_columns:
        pf_col = full_columns.index("PF")
        aci_col = full_columns.index("ACI")
        ws.conditional_format(
            1,
            pf_col,
            last_row,
            pf_col,
            {
                "type": "3_color_scale",
                "min_color": "#63BE7B",
                "mid_type": "percentile",
                "mid_value": 50,
                "mid_color": "#FFEB84",
                "max_color": "#F8696B",
            },
        )
        ws.conditional_format(
            1,
            aci_col,
            last_row,
            aci_col,
            {
                "type": "3_color_scale",
                "min_color": "#F8696B",
                "mid_type": "percentile",
                "mid_value": 50,
                "mid_color": "#FFEB84",
                "max_color": "#63BE7B",
            },
        )


def write_summary_sheet(
    ws: Any,
    fmts: dict[str, Any],
    priority_sheet_name: str,
    full_columns: list[str],
    rows: list[AssetRow],
) -> None:
    """Write Sheet 2: Summary Statistics with live Excel formulas."""
    total = len(rows)
    last_row = total + 1  # 1-based last data row on the priority sheet

    ws.set_column(0, 0, 36)
    ws.set_column(1, 5, 18)

    def col_range(name: str) -> str:
        idx_zero = full_columns.index(name)
        letter = _xl_col_letter(idx_zero)
        return f"'{priority_sheet_name}'!{letter}2:{letter}{last_row}"

    pf_range = col_range("PF") if "PF" in full_columns else ""

    ws.write_string(0, 0, "Master Priority Output", fmts["title"])
    ws.write_string(1, 0, "Summary Statistics", fmts["subtitle"])
    ws.write_string(
        2,
        0,
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        fmts["italic_grey"],
    )
    ws.write_string(3, 0, f"Total Assets: {total}", fmts["bold_small"])

    ws.write_string(5, 0, "Factor Statistics", fmts["section"])
    factor_headers = ("Factor", "Average", "Median", "Min", "Max", "Count")
    for c, h in enumerate(factor_headers):
        ws.write_string(6, c, h, fmts["header"])

    factor_names = ["ACI", "CF", "FLF", "RF", "PF"]
    row_cursor = 7
    factor_row_count = 0
    for factor in factor_names:
        if factor not in full_columns:
            continue
        rng = col_range(factor)
        ws.write_string(row_cursor, 0, factor, fmts["summary_label"])
        ws.write_formula(row_cursor, 1, f"=IFERROR(AVERAGE({rng}),0)", fmts["body_num"])
        ws.write_formula(row_cursor, 2, f"=IFERROR(MEDIAN({rng}),0)", fmts["body_num"])
        ws.write_formula(row_cursor, 3, f"=IFERROR(MIN({rng}),0)", fmts["body_num"])
        ws.write_formula(row_cursor, 4, f"=IFERROR(MAX({rng}),0)", fmts["body_num"])
        ws.write_formula(row_cursor, 5, f"=COUNT({rng})", fmts["body"])
        row_cursor += 1
        factor_row_count += 1

    row_cursor += 1
    ws.write_string(row_cursor, 0, "Asset Count by Type", fmts["section"])
    row_cursor += 1
    for c, h in enumerate(("Asset Type", "Count", "% Share")):
        ws.write_string(row_cursor, c, h, fmts["header"])
    row_cursor += 1

    type_counter = Counter(
        _to_str(r.canonical.get("Asset_Type"))
        for r in rows
        if r.canonical.get("Asset_Type")
    )
    type_row_count = 0
    if "Asset_Type" in full_columns:
        type_col_letter = _xl_col_letter(full_columns.index("Asset_Type"))
        type_range = (
            f"'{priority_sheet_name}'!{type_col_letter}2:{type_col_letter}{last_row}"
        )
        for asset_type, _count in sorted(type_counter.items(), key=lambda x: -x[1]):
            excel_row_1based = row_cursor + 1
            ws.write_string(row_cursor, 0, asset_type, fmts["body"])
            ws.write_formula(
                row_cursor,
                1,
                f'=COUNTIF({type_range},"{_escape_for_formula(asset_type)}")',
                fmts["body"],
            )
            ws.write_formula(
                row_cursor,
                2,
                f"=IFERROR(B{excel_row_1based}/COUNTA({type_range}),0)",
                fmts["body_pct"],
            )
            row_cursor += 1
            type_row_count += 1

    if pf_range:
        row_cursor += 1
        ws.write_string(row_cursor, 0, "Priority Bands (by PF)", fmts["section"])
        row_cursor += 1
        for c, h in enumerate(("Band", "Threshold", "Count")):
            ws.write_string(row_cursor, c, h, fmts["header"])
        row_cursor += 1

        bands = [
            ("Critical (Top 10%)", f"=PERCENTILE({pf_range},0.90)"),
            ("High (10-25%)", f"=PERCENTILE({pf_range},0.75)"),
            ("Medium (25-50%)", f"=PERCENTILE({pf_range},0.50)"),
            ("Low (Bottom 50%)", ""),
        ]
        band_start_row_0based = row_cursor
        for i, (label, formula) in enumerate(bands):
            r = band_start_row_0based + i
            r_1based = r + 1
            upper_1based = band_start_row_0based + (i - 1) + 1
            if i == 0:
                count_formula = f'=COUNTIF({pf_range},">="&B{r_1based})'
            elif i in (1, 2):
                count_formula = (
                    f'=COUNTIFS({pf_range},">="&B{r_1based},'
                    f'{pf_range},"<"&B{upper_1based})'
                )
            else:
                count_formula = f'=COUNTIF({pf_range},"<"&B{upper_1based})'
            ws.write_string(r, 0, label, fmts["body"])
            if formula:
                ws.write_formula(r, 1, formula, fmts["body_num"])
            else:
                ws.write_blank(r, 1, None, fmts["body_num"])
            ws.write_formula(r, 2, count_formula, fmts["body"])
        row_cursor = band_start_row_0based + len(bands)

    row_cursor += 1
    ws.write_string(row_cursor, 0, "Top 10 Highest-Priority Assets", fmts["section"])
    row_cursor += 1
    top_headers = ("Rank", "Asset_ID", "Asset_Type", "Road_Name_EN", "ACI", "PF")
    for c, h in enumerate(top_headers):
        ws.write_string(row_cursor, c, h, fmts["header"])
    row_cursor += 1

    top_field_letters = {
        name: _xl_col_letter(full_columns.index(name))
        for name in ("Asset_ID", "Asset_Type", "Road_Name_EN", "ACI", "PF")
        if name in full_columns
    }
    for rank in range(1, 11):
        sheet_row = rank + 1
        if rank > total:
            break
        ws.write_number(row_cursor, 0, rank, fmts["body_center"])
        for col_offset, name in enumerate(
            ("Asset_ID", "Asset_Type", "Road_Name_EN", "ACI", "PF"), start=1
        ):
            fmt = fmts["body_num"] if name in {"ACI", "PF"} else fmts["body"]
            if name in top_field_letters:
                ref = f"'{priority_sheet_name}'!{top_field_letters[name]}{sheet_row}"
                ws.write_formula(row_cursor, col_offset, f"={ref}", fmt)
            else:
                ws.write_blank(row_cursor, col_offset, None, fmt)
        row_cursor += 1


def write_source_mapping_sheet(
    ws: Any,
    fmts: dict[str, Any],
    extracts: list[SheetExtract],
) -> None:
    """Write Sheet 3: Source Mapping with per-file and per-sheet counts."""
    ws.set_column(0, 0, 36)
    ws.set_column(1, 1, 32)
    ws.set_column(2, 2, 60)
    ws.set_column(3, 3, 14)

    ws.write_string(0, 0, "Source Mapping", fmts["title"])
    ws.write_string(
        1,
        0,
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        fmts["italic_grey"],
    )

    row = 3
    headers = ("Source File", "Source Sheet", "Asset Types", "Row Count")
    for c, h in enumerate(headers):
        ws.write_string(row, c, h, fmts["header"])
    row += 1

    file_totals: Counter[str] = Counter()
    for extract in sorted(extracts, key=lambda e: (e.file_name, e.sheet_name)):
        type_counts: Counter[str] = Counter()
        for r in extract.rows:
            t = _to_str(r.canonical.get("Asset_Type"))
            if t:
                type_counts[t] += 1
        types_text = ", ".join(
            f"{name} ({count})" for name, count in type_counts.most_common()
        )
        ws.write_string(row, 0, extract.file_name, fmts["body"])
        ws.write_string(row, 1, extract.sheet_name, fmts["body"])
        ws.write_string(row, 2, types_text, fmts["body"])
        ws.write_number(row, 3, len(extract.rows), fmts["body"])
        file_totals[extract.file_name] += len(extract.rows)
        row += 1

    row += 1
    ws.write_string(row, 0, "Per-File Totals", fmts["section"])
    row += 1
    for c, h in enumerate(("Source File", "Total Rows")):
        ws.write_string(row, c, h, fmts["header"])
    row += 1
    for file_name, total in file_totals.most_common():
        ws.write_string(row, 0, file_name, fmts["body"])
        ws.write_number(row, 1, total, fmts["body"])
        row += 1


# ---------------------------------------------------------------------------
# Discovery + orchestration
# ---------------------------------------------------------------------------


def discover_input_files(input_dir: Path) -> list[Path]:
    """Find all .xlsx files in ``input_dir`` excluding our own output."""
    candidates = sorted(input_dir.glob("*.xlsx"))
    return [
        p
        for p in candidates
        if not p.name.startswith("~$")
        and p.name.lower() != "master_priority_output.xlsx"
    ]


def process_file(path: Path) -> list[SheetExtract]:
    """Open one workbook and extract every inventory sheet."""
    _log(f"  Reading: {path.name}")
    extracts: list[SheetExtract] = []
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        _log(f"    ! Could not open: {exc}")
        return extracts
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            t0 = time.perf_counter()
            try:
                extract = extract_sheet(ws, path.name, sheet_name)
            except Exception as exc:
                _log(f"    ! Sheet '{sheet_name}' failed: {exc}")
                continue
            dt = time.perf_counter() - t0
            if extract is None or not extract.rows:
                if dt > 0.5:
                    _log(f"    . {sheet_name}: skipped (no inventory data, {dt:.1f}s)")
                continue
            _log(f"    - {sheet_name}: {len(extract.rows)} rows ({dt:.1f}s)")
            extracts.append(extract)
    finally:
        wb.close()
    return extracts


def generate_master_priority(
    input_dir: Path,
    output_path: Path,
    explicit_files: Optional[list[Path]] = None,
) -> Path:
    """End-to-end pipeline. Returns the output path."""
    _log("=" * 70)
    _log("Master Priority Output Generator")
    _log("=" * 70)

    files = explicit_files if explicit_files else discover_input_files(input_dir)
    if not files:
        raise FileNotFoundError(f"No .xlsx files found in {input_dir}")

    _log(f"Input directory: {input_dir}")
    _log(f"Discovered files: {len(files)}")
    for p in files:
        _log(f"  * {p.name}")

    all_extracts: list[SheetExtract] = []
    for path in files:
        if not path.exists():
            _log(f"  ! Missing: {path}")
            continue
        all_extracts.extend(process_file(path))

    if not all_extracts:
        raise RuntimeError("No inventory rows extracted from any input file.")

    _log("\nConsolidating rows...")
    t0 = time.perf_counter()
    extra_columns, all_rows = consolidate(all_extracts)
    _log(f"Consolidation: {len(all_rows)} rows in {time.perf_counter() - t0:.1f}s")
    computed_pf = sum(1 for r in all_rows if r.canonical.get("PF_Computed"))
    missing_pf = sum(1 for r in all_rows if _to_float(r.canonical.get("PF")) is None)

    _log(f"\nTotal rows consolidated: {len(all_rows)}")
    _log(f"  PF from source:               {len(all_rows) - computed_pf - missing_pf}")
    _log(f"  PF computed by us:            {computed_pf}")
    _log(f"  PF missing (no PF/CF/RF/FLF): {missing_pf}")

    full_columns = _deduplicate_columns(PRIORITY_COLUMNS + extra_columns)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    _log("\nWriting output workbook (xlsxwriter)...")
    wb = xlsxwriter.Workbook(
        str(output_path),
        {"constant_memory": True, "default_date_format": "yyyy-mm-dd"},
    )
    try:
        fmts = _build_formats(wb)

        ws1 = wb.add_worksheet("Priority Ranking")
        t0 = time.perf_counter()
        write_priority_sheet(ws1, fmts, full_columns, all_rows)
        _log(f"  Sheet 1 (Priority Ranking): {time.perf_counter() - t0:.1f}s")

        ws2 = wb.add_worksheet("Summary Statistics")
        t0 = time.perf_counter()
        write_summary_sheet(ws2, fmts, "Priority Ranking", full_columns, all_rows)
        _log(f"  Sheet 2 (Summary Statistics): {time.perf_counter() - t0:.1f}s")

        ws3 = wb.add_worksheet("Source Mapping")
        t0 = time.perf_counter()
        write_source_mapping_sheet(ws3, fmts, all_extracts)
        _log(f"  Sheet 3 (Source Mapping): {time.perf_counter() - t0:.1f}s")
    finally:
        t0 = time.perf_counter()
        wb.close()
        _log(f"  Close + flush: {time.perf_counter() - t0:.1f}s")

    _log(f"\nOutput written: {output_path}")
    _log("=" * 70)
    return output_path


def main(argv: Optional[list[str]] = None) -> int:
    parser = argparse.ArgumentParser(
        description="Consolidate multi-file asset inventory into a "
        "Priority-Factor-ranked master Excel file."
    )
    parser.add_argument(
        "--input-dir",
        type=Path,
        default=DEFAULT_INPUT_DIR,
        help=f"Directory containing input .xlsx files (default: {DEFAULT_INPUT_DIR})",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT_FILE,
        help=f"Output .xlsx path (default: {DEFAULT_OUTPUT_FILE})",
    )
    parser.add_argument(
        "files",
        nargs="*",
        type=Path,
        help="Optional explicit input files (overrides --input-dir discovery)",
    )
    args = parser.parse_args(argv)

    explicit = [p.resolve() for p in args.files] if args.files else None
    output_path = args.output.resolve()
    input_dir = args.input_dir.resolve()

    generate_master_priority(input_dir, output_path, explicit_files=explicit)
    return 0


if __name__ == "__main__":
    sys.exit(main())
