"""
Extract lookup tables from RTA Excel workbooks to JSON format.

This script reads the three source Excel files and exports:
1. Defect mappings per asset type (from ACI Sheet)
2. FLF deterioration equations (from Priority Factor workbook)
3. RF risk scores (from Priority Factor workbook)

Run once to generate all JSON files for the domain package.
"""

import json
import re
from pathlib import Path
from openpyxl import load_workbook


BASE_DIR = Path(__file__).parent.parent
EXCEL_DIR = BASE_DIR.parent
LOOKUPS_DIR = BASE_DIR / "domain" / "lookups"
DEFECTS_DIR = LOOKUPS_DIR / "defects"

PRIORITY_FACTOR_FILE = EXCEL_DIR / "1. Calculations for Priority Index.xlsx"
ACI_SHEET_FILE = EXCEL_DIR / "RoW Assets_ACI Sheet_20-05-2025.xlsx"
MASTER_EXCEL_FILE = EXCEL_DIR / "MASTER_EXCEL_E11.xlsx"


def extract_rf_scores():
    """Extract Risk Factor scores from Priority Factor workbook."""
    print("Extracting RF scores...")
    wb = load_workbook(PRIORITY_FACTOR_FILE, data_only=True)
    ws = wb["Risk Factor"]
    
    rf_scores = {}
    for row in ws.iter_rows(min_row=2, max_row=35, max_col=5):
        sn = row[0].value
        name = row[1].value
        priority = row[2].value
        risk_score = row[3].value
        
        if name and sn and risk_score:
            asset_key = normalize_asset_name(str(name))
            rf_scores[asset_key] = {
                "name": str(name),
                "priority_level": str(priority) if priority else "M",
                "risk_score": int(risk_score)
            }
    
    wb.close()
    
    output_path = LOOKUPS_DIR / "rf_scores.json"
    with open(output_path, "w") as f:
        json.dump(rf_scores, f, indent=2)
    print(f"  Saved {len(rf_scores)} RF scores to {output_path}")
    return rf_scores


def extract_flf_equations():
    """Extract FLF deterioration equations from Priority Factor workbook."""
    print("Extracting FLF equations...")
    wb = load_workbook(PRIORITY_FACTOR_FILE, data_only=False)
    ws = wb["Functional Life Factor "]  # Note trailing space
    
    flf_equations = {}
    for row in ws.iter_rows(min_row=2, max_row=35, max_col=15):
        sn = row[0].value
        name = row[1].value
        equation_text = row[6].value  # Column G
        total_life = row[8].value     # Column I
        
        if name and sn and equation_text and total_life:
            asset_key = normalize_asset_name(str(name))
            intercept, slope = parse_deterioration_equation(str(equation_text))
            
            if intercept is not None and slope is not None:
                flf_equations[asset_key] = {
                    "name": str(name),
                    "equation": str(equation_text),
                    "intercept": intercept,
                    "slope": slope,
                    "total_life": float(total_life)
                }
    
    wb.close()
    
    output_path = LOOKUPS_DIR / "flf_equations.json"
    with open(output_path, "w") as f:
        json.dump(flf_equations, f, indent=2)
    print(f"  Saved {len(flf_equations)} FLF equations to {output_path}")
    return flf_equations


def parse_deterioration_equation(equation: str) -> tuple:
    """
    Parse equation like '(ACI-99.435)/-3.0273' into intercept and slope.
    Returns (intercept, slope) or (None, None) if parsing fails.
    """
    pattern = r"\(ACI[- ]*([0-9.]+)\)/([- ]*[0-9.]+)"
    match = re.search(pattern, equation.replace(" ", ""))
    if match:
        intercept = float(match.group(1))
        slope = float(match.group(2))
        return intercept, slope
    return None, None


def normalize_asset_name(name: str) -> str:
    """Convert asset name to consistent key format."""
    name = name.upper().strip()
    name = re.sub(r"[^A-Z0-9]+", "_", name)
    name = re.sub(r"_+", "_", name)
    name = name.strip("_")
    return name


def extract_defect_lookups():
    """Extract defect-to-score mappings from ACI Sheet workbook."""
    print("Extracting defect lookups...")
    wb = load_workbook(ACI_SHEET_FILE, data_only=True)
    
    extracted_count = 0
    for sheet_name in wb.sheetnames:
        if sheet_name in ["ACI_Calculation Sheet", "Summary"]:
            continue
        
        ws = wb[sheet_name]
        defect_data = extract_sheet_defects(ws, sheet_name)
        
        if defect_data:
            asset_key = normalize_asset_name(sheet_name)
            output_path = DEFECTS_DIR / f"{asset_key.lower()}.json"
            with open(output_path, "w") as f:
                json.dump(defect_data, f, indent=2)
            extracted_count += 1
    
    wb.close()
    print(f"  Saved {extracted_count} defect lookup files to {DEFECTS_DIR}")


def extract_sheet_defects(ws, sheet_name: str) -> dict:
    """Extract defect mappings from a single asset sheet."""
    defect_data = {
        "asset_type": normalize_asset_name(sheet_name),
        "sheet_name": sheet_name,
        "defect_categories": [],
        "aci_formula": "",
        "weights": {}
    }
    
    functional_defects = []
    appearance_defects = []
    other_defects = {}
    
    max_row = min(ws.max_row, 30)
    max_col = min(ws.max_column, 25)
    
    for col in range(1, max_col + 1):
        header = ws.cell(1, col).value
        if not header:
            continue
        header_lower = str(header).lower()
        
        if "functional" in header_lower and "defect" in header_lower:
            defects = extract_column_defects(ws, col, max_row)
            if defects:
                functional_defects.extend(defects)
        elif "appearance" in header_lower and "defect" in header_lower:
            defects = extract_column_defects(ws, col, max_row)
            if defects:
                appearance_defects.extend(defects)
        elif "defect" in header_lower:
            category = infer_category_name(header)
            defects = extract_column_defects(ws, col, max_row)
            if defects:
                other_defects[category] = defects
    
    lookup_tables = extract_lookup_tables(ws, max_row, max_col)
    
    if functional_defects or lookup_tables.get("functional"):
        defect_data["defect_categories"].append({
            "name": "functional",
            "max_score": 80,
            "defects": functional_defects or lookup_tables.get("functional", [])
        })
    
    if appearance_defects or lookup_tables.get("appearance"):
        defect_data["defect_categories"].append({
            "name": "appearance", 
            "max_score": 20,
            "defects": appearance_defects or lookup_tables.get("appearance", [])
        })
    
    for category, defects in other_defects.items():
        if defects:
            defect_data["defect_categories"].append({
                "name": category,
                "defects": defects
            })
    
    for category, defects in lookup_tables.items():
        if category not in ["functional", "appearance"] and defects:
            existing = [c for c in defect_data["defect_categories"] if c["name"] == category]
            if not existing:
                defect_data["defect_categories"].append({
                    "name": category,
                    "defects": defects
                })
    
    defect_data["aci_formula"] = infer_aci_formula(ws, sheet_name)
    defect_data["weights"] = infer_weights(ws, sheet_name)
    
    return defect_data if defect_data["defect_categories"] else None


def extract_column_defects(ws, col: int, max_row: int) -> list:
    """Extract defect text entries from a column."""
    defects = []
    seen = set()
    
    for row in range(2, max_row + 1):
        value = ws.cell(row, col).value
        if value and isinstance(value, str):
            text = value.strip()
            if text and text not in seen and not text.startswith("Select"):
                seen.add(text)
                defects.append({"text": text, "score": None})
    
    return defects


def extract_lookup_tables(ws, max_row: int, max_col: int) -> dict:
    """Extract lookup tables (defect text -> score mappings) from sheet."""
    lookups = {}
    
    for col in range(6, max_col + 1):
        defect_texts = []
        scores = []
        
        for row in range(2, max_row + 1):
            text_val = ws.cell(row, col).value
            score_col = col + 1 if col + 1 <= max_col else col + 2
            score_val = ws.cell(row, score_col).value if score_col <= max_col else None
            
            if text_val and isinstance(text_val, str) and not str(text_val).startswith("Select"):
                text = str(text_val).strip()
                if text and "defect" not in text.lower()[:20]:
                    score = None
                    if score_val is not None:
                        try:
                            score = int(float(score_val))
                        except (ValueError, TypeError):
                            pass
                    defect_texts.append({"text": text, "score": score})
        
        if defect_texts:
            header = ws.cell(1, col).value or ws.cell(2, col).value or f"category_{col}"
            category = infer_category_name(str(header))
            if category not in lookups:
                lookups[category] = defect_texts
            else:
                lookups[category].extend(defect_texts)
    
    return lookups


def infer_category_name(header: str) -> str:
    """Infer defect category from header text."""
    header_lower = header.lower()
    if "functional" in header_lower or "function" in header_lower:
        return "functional"
    if "appearance" in header_lower:
        return "appearance"
    if "foundation" in header_lower:
        return "foundation"
    if "structure" in header_lower:
        return "structure"
    if "visibility" in header_lower:
        return "visibility"
    if "presence" in header_lower:
        return "presence"
    if "blockage" in header_lower:
        return "blockage"
    if "concrete" in header_lower:
        return "concrete_damage"
    if "roof" in header_lower:
        return "roof"
    if "column" in header_lower:
        return "column"
    if "pole" in header_lower:
        return "pole"
    if "signal" in header_lower or "light" in header_lower:
        return "signal_head"
    return "other"


def infer_aci_formula(ws, sheet_name: str) -> str:
    """Infer ACI formula from sheet structure."""
    sheet_upper = sheet_name.upper()
    
    if sheet_upper in ["GANTRY"]:
        return "0.4*F + 0.3*SignScore + 0.3*(V+P)"
    if sheet_upper in ["TRAFIC SIGNAL", "TRAFFIC SIGNAL"]:
        return "0.2*F + 0.2*(FC+A) + 0.6*SignalHead"
    if sheet_upper in ["STREETLIGHT"]:
        return "0.2*F + 0.2*(FC+A) + 0.3*LightingArm + 0.3*Lamp"
    if sheet_upper in ["PERGOLA", "PUBLIC SHED", "PARKING TENTS"]:
        return "0.4*F + 0.3*(FC+A) + 0.3*Roof"
    if sheet_upper in ["PEDESTRIAN CROSSING"]:
        return "0.5*Structural + 0.5*(V+P)"
    if sheet_upper in ["ROAD_SIGN_FACE AND POLE"]:
        return "0.9*FaceACI + 0.1*PoleACI"
    if sheet_upper in ["BOLLARDS"]:
        return "F*0.7 + A*0.3"
    if sheet_upper in ["ROAD MARKING", "ROAD MARKING SYMBOL"]:
        return "V*0.7 + P*0.3"
    if sheet_upper in ["BARRIER"]:
        return "F*0.8 + A*0.2"
    if sheet_upper in ["ROAD STUD", "FOOTPATH", "CAMEL GRID", "CONTROL CABINET", 
                       "ROADSIDE CAMERA", "DECORATIVE FURNITURE", "LANDSCAPE",
                       "PARKING", "LABAY", "SHOULDER", "ISLAND"]:
        return "F"
    return "F + A"


def infer_weights(ws, sheet_name: str) -> dict:
    """Infer component weights from sheet structure."""
    sheet_upper = sheet_name.upper()
    
    if sheet_upper in ["GANTRY"]:
        return {"foundation": 0.4, "sign_score": 0.3, "visibility_presence": 0.3}
    if sheet_upper in ["TRAFIC SIGNAL", "TRAFFIC SIGNAL"]:
        return {"foundation": 0.2, "pole": 0.2, "signal_head": 0.6}
    if sheet_upper in ["STREETLIGHT"]:
        return {"foundation": 0.2, "pole": 0.2, "lighting_arm": 0.3, "lamp": 0.3}
    if sheet_upper in ["PERGOLA", "PUBLIC SHED", "PARKING TENTS"]:
        return {"foundation": 0.4, "column": 0.3, "roof": 0.3}
    if sheet_upper in ["PEDESTRIAN CROSSING"]:
        return {"structural": 0.5, "visibility_presence": 0.5}
    if sheet_upper in ["ROAD_SIGN_FACE AND POLE"]:
        return {"face": 0.9, "pole": 0.1}
    if sheet_upper in ["BOLLARDS"]:
        return {"functional": 0.7, "appearance": 0.3}
    if sheet_upper in ["ROAD MARKING", "ROAD MARKING SYMBOL"]:
        return {"visibility": 0.7, "presence": 0.3}
    if sheet_upper in ["BARRIER"]:
        return {"functional": 0.8, "appearance": 0.2}
    if sheet_upper in ["ROAD STUD", "FOOTPATH", "CAMEL GRID", "CONTROL CABINET",
                       "ROADSIDE CAMERA", "DECORATIVE FURNITURE", "LANDSCAPE",
                       "PARKING", "LABAY", "SHOULDER", "ISLAND"]:
        return {"functional": 1.0}
    return {"functional": 0.8, "appearance": 0.2}


def extract_asset_type_mapping():
    """Create a mapping of all asset types with their calculation categories."""
    print("Creating asset type mapping...")
    
    asset_types = {
        "simple_f_a": [
            "MANHOLE", "GULLY", "CURBSTONE", "GUARDRAIL", "FENCE", "HUMP",
            "BENCH", "CYCLE_RACK", "WOODEN_DECK", "JOGGING_TRACK", "ROAD_SIGN_POLE",
            "DRAINAGE_POINT", "CRASH_CUSHION_END_TERMINAL"
        ],
        "functional_only": [
            "ROAD_STUD", "FOOTPATH", "CAMEL_GRID", "CONTROL_CABINET",
            "ROADSIDE_CAMERA", "DECORATIVE_FURNITURE", "LANDSCAPE",
            "PARKING", "LABAY", "SHOULDER", "ISLAND"
        ],
        "weighted_70_30": [
            "BOLLARDS", "ROAD_MARKING", "ROAD_MARKING_SYMBOL"
        ],
        "weighted_80_20": [
            "BARRIER"
        ],
        "complex": [
            "GANTRY", "TRAFIC_SIGNAL", "STREETLIGHT", "PERGOLA",
            "PUBLIC_SHED", "PARKING_TENTS", "PEDESTRIAN_CROSSING",
            "ROAD_SIGN_FACE_AND_POLE"
        ]
    }
    
    output_path = LOOKUPS_DIR / "asset_types.json"
    with open(output_path, "w") as f:
        json.dump(asset_types, f, indent=2)
    print(f"  Saved asset type mapping to {output_path}")
    return asset_types


def main():
    """Run all extraction steps."""
    print("=" * 60)
    print("RTA Excel Lookup Extraction")
    print("=" * 60)
    
    LOOKUPS_DIR.mkdir(parents=True, exist_ok=True)
    DEFECTS_DIR.mkdir(parents=True, exist_ok=True)
    
    print(f"\nSource files:")
    print(f"  Priority Factor: {PRIORITY_FACTOR_FILE}")
    print(f"  ACI Sheet: {ACI_SHEET_FILE}")
    print(f"  Master Excel: {MASTER_EXCEL_FILE}")
    print(f"\nOutput directory: {LOOKUPS_DIR}")
    print()
    
    rf_scores = extract_rf_scores()
    flf_equations = extract_flf_equations()
    extract_defect_lookups()
    extract_asset_type_mapping()
    
    print()
    print("=" * 60)
    print("Extraction complete!")
    print("=" * 60)


if __name__ == "__main__":
    main()
