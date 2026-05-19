"""
Excel Test Report Generator for RTA Asset Management.

Generates comprehensive Excel reports from pytest test results,
suitable for presenting to supervising engineers.
"""

from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import PieChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
HEADER_FILL = PatternFill(start_color="1E3C72", end_color="1E3C72", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="1E3C72")
SUBTITLE_FONT = Font(bold=True, size=12, color="2A5298")
PASS_FONT = Font(bold=True, color="006100")
FAIL_FONT = Font(bold=True, color="9C0006")

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

MODULE_NAMES = {
    "test_aci": "ACI Tests",
    "test_cf": "CF Tests",
    "test_flf": "FLF Tests",
    "test_rf": "RF Tests",
    "test_pf": "PF Tests",
}

MODULE_DESCRIPTIONS = {
    "test_aci": "Asset Condition Index - measures overall asset condition (0-100)",
    "test_cf": "Condition Factor - priority boost based on condition (0-100)",
    "test_flf": "Functional Life Factor - percentage of useful life consumed (0-100)",
    "test_rf": "Risk Factor - inherent risk/importance of asset type (10-100)",
    "test_pf": "Priority Factor - aggregated score for maintenance prioritization",
}

MODULE_FORMULAS = {
    "test_aci": "ACI = Functional Score + Appearance Score (varies by asset category)",
    "test_cf": "CF = IF(ACI >= 80, 0, (80 - ACI) / 80 * 100)",
    "test_flf": "FLF = IF(ACI=0, 100, IF(ACI>=80, 0, (ConsumedLife/TotalLife)*100))",
    "test_rf": "RF = 10 + ((Score - MIN) / (MAX - MIN)) * 90",
    "test_pf": "PF = CF * 0.6 + RF * 0.2 + FLF * 0.2",
}


class ExcelTestReporter:
    """Generates Excel reports from pytest test results."""

    def __init__(self, output_dir: Path = None):
        self.output_dir = output_dir or Path.cwd()
        self.results: list[dict[str, Any]] = []
        self.start_time: datetime = None
        self.end_time: datetime = None

    def start_session(self):
        """Mark the start of a test session."""
        self.start_time = datetime.now()
        self.results = []

    def add_result(self, result: dict[str, Any]):
        """Add a test result to the collection."""
        self.results.append(result)

    def end_session(self):
        """Mark the end of a test session."""
        self.end_time = datetime.now()

    def generate_report(self) -> Path:
        """Generate the Excel report and return the file path."""
        timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
        filename = f"test_report_{timestamp}.xlsx"
        output_path = self.output_dir / filename

        wb = Workbook()
        
        self._create_summary_sheet(wb)
        
        modules = self._group_by_module()
        for module_name in ["test_aci", "test_cf", "test_flf", "test_rf", "test_pf"]:
            if module_name in modules:
                self._create_module_sheet(wb, module_name, modules[module_name])

        wb.save(output_path)
        return output_path

    def _group_by_module(self) -> dict[str, list[dict]]:
        """Group test results by module."""
        modules = {}
        for result in self.results:
            module = result.get("module", "unknown")
            module_key = module.split(".")[-1] if "." in module else module
            if module_key not in modules:
                modules[module_key] = []
            modules[module_key].append(result)
        return modules

    def _create_summary_sheet(self, wb: Workbook):
        """Create the summary sheet with statistics and chart."""
        ws = wb.active
        ws.title = "Summary"

        total = len(self.results)
        passed = sum(1 for r in self.results if r.get("status") == "PASS")
        failed = sum(1 for r in self.results if r.get("status") == "FAIL")
        skipped = sum(1 for r in self.results if r.get("status") == "SKIP")
        pass_rate = (passed / total * 100) if total > 0 else 0
        total_duration = sum(r.get("duration_ms", 0) for r in self.results)

        ws["A1"] = "RTA Asset Management"
        ws["A1"].font = TITLE_FONT
        ws["A2"] = "Test Execution Report"
        ws["A2"].font = SUBTITLE_FONT

        ws["A4"] = "Generated:"
        ws["B4"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws["A5"] = "Test Duration:"
        ws["B5"] = f"{total_duration:.2f} ms"

        ws["A7"] = "Test Results Summary"
        ws["A7"].font = SUBTITLE_FONT

        headers = ["Metric", "Value"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=8, column=col, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")

        stats = [
            ("Total Tests", total),
            ("Passed", passed),
            ("Failed", failed),
            ("Skipped", skipped),
            ("Pass Rate", f"{pass_rate:.1f}%"),
        ]

        for row_idx, (metric, value) in enumerate(stats, 9):
            ws.cell(row=row_idx, column=1, value=metric).border = THIN_BORDER
            value_cell = ws.cell(row=row_idx, column=2, value=value)
            value_cell.border = THIN_BORDER
            value_cell.alignment = Alignment(horizontal="center")
            
            if metric == "Passed":
                value_cell.fill = PASS_FILL
                value_cell.font = PASS_FONT
            elif metric == "Failed" and isinstance(value, (int, float)) and value > 0:
                value_cell.fill = FAIL_FILL
                value_cell.font = FAIL_FONT

        ws["A15"] = "Results by Module"
        ws["A15"].font = SUBTITLE_FONT

        module_headers = ["Module", "Tests", "Passed", "Failed", "Pass Rate"]
        for col, header in enumerate(module_headers, 1):
            cell = ws.cell(row=16, column=col, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")

        modules = self._group_by_module()
        row = 17
        for module_name in ["test_aci", "test_cf", "test_flf", "test_rf", "test_pf"]:
            if module_name in modules:
                results = modules[module_name]
                m_total = len(results)
                m_passed = sum(1 for r in results if r.get("status") == "PASS")
                m_failed = sum(1 for r in results if r.get("status") == "FAIL")
                m_rate = (m_passed / m_total * 100) if m_total > 0 else 0

                display_name = MODULE_NAMES.get(module_name, module_name)
                
                ws.cell(row=row, column=1, value=display_name).border = THIN_BORDER
                ws.cell(row=row, column=2, value=m_total).border = THIN_BORDER
                
                passed_cell = ws.cell(row=row, column=3, value=m_passed)
                passed_cell.border = THIN_BORDER
                passed_cell.fill = PASS_FILL
                
                failed_cell = ws.cell(row=row, column=4, value=m_failed)
                failed_cell.border = THIN_BORDER
                if m_failed > 0:
                    failed_cell.fill = FAIL_FILL
                
                ws.cell(row=row, column=5, value=f"{m_rate:.1f}%").border = THIN_BORDER
                
                for col in range(1, 6):
                    ws.cell(row=row, column=col).alignment = Alignment(horizontal="center")
                
                row += 1

        if passed > 0 or failed > 0:
            ws["D4"] = "Status"
            ws["E4"] = "Count"
            ws["D5"] = "Passed"
            ws["E5"] = float(passed)
            ws["D6"] = "Failed"
            ws["E6"] = float(failed)

            try:
                chart = PieChart()
                chart.title = "Test Results"
                labels = Reference(ws, min_col=4, min_row=5, max_row=6)
                data = Reference(ws, min_col=5, min_row=5, max_row=6)
                chart.add_data(data)
                chart.set_categories(labels)
                chart.width = 10
                chart.height = 8
                ws.add_chart(chart, "G4")
            except Exception:
                pass

        self._auto_width_columns(ws)

    def _create_module_sheet(self, wb: Workbook, module_name: str, results: list[dict]):
        """Create a sheet for a specific test module."""
        display_name = MODULE_NAMES.get(module_name, module_name)
        ws = wb.create_sheet(title=display_name)

        ws["A1"] = display_name
        ws["A1"].font = TITLE_FONT
        
        description = MODULE_DESCRIPTIONS.get(module_name, "")
        ws["A2"] = description
        ws["A2"].font = Font(italic=True, color="666666")
        
        ws["A3"] = "Formula:"
        ws["A3"].font = Font(bold=True)
        ws["B3"] = MODULE_FORMULAS.get(module_name, "N/A")
        ws["B3"].font = Font(name="Consolas", size=10)

        headers = [
            "Test Name",
            "Description",
            "Asset Type(s)",
            "Input Values",
            "Expected Output",
            "Status",
            "Duration (ms)",
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        row = 6
        for result in results:
            test_name = result.get("name", "")
            docstring = result.get("docstring", "") or ""
            params = result.get("params", {})
            status = result.get("status", "UNKNOWN")
            duration = result.get("duration_ms", 0)

            asset_types = self._extract_asset_types(test_name, params)
            input_values = self._format_params(params)
            expected = self._extract_expected(docstring)

            ws.cell(row=row, column=1, value=test_name).border = THIN_BORDER
            
            desc_cell = ws.cell(row=row, column=2, value=docstring.strip() if docstring else "")
            desc_cell.border = THIN_BORDER
            desc_cell.alignment = Alignment(wrap_text=True)
            
            ws.cell(row=row, column=3, value=asset_types).border = THIN_BORDER
            
            input_cell = ws.cell(row=row, column=4, value=input_values)
            input_cell.border = THIN_BORDER
            input_cell.alignment = Alignment(wrap_text=True)
            
            ws.cell(row=row, column=5, value=expected).border = THIN_BORDER

            status_cell = ws.cell(row=row, column=6, value=status)
            status_cell.border = THIN_BORDER
            status_cell.alignment = Alignment(horizontal="center")
            if status == "PASS":
                status_cell.fill = PASS_FILL
                status_cell.font = PASS_FONT
            elif status == "FAIL":
                status_cell.fill = FAIL_FILL
                status_cell.font = FAIL_FONT

            duration_cell = ws.cell(row=row, column=7, value=f"{duration:.2f}")
            duration_cell.border = THIN_BORDER
            duration_cell.alignment = Alignment(horizontal="right")

            row += 1

        self._auto_width_columns(ws, max_width=50)

    def _extract_asset_types(self, test_name: str, params: dict) -> str:
        """Extract asset types from test name or parameters."""
        if "asset_type" in params:
            return str(params["asset_type"])
        
        asset_keywords = [
            "MANHOLE", "GULLY", "FOOTPATH", "BOLLARDS", "BARRIER", "GANTRY",
            "TRAFFIC_SIGNAL", "STREETLIGHT", "BENCHES", "ROAD_SIGN", "GUARDRAIL",
            "FENCE", "CRASH_CUSHION", "ROAD_MARKING", "ROAD_STUDS", "PEDESTRIAN",
        ]
        
        found = []
        name_upper = test_name.upper()
        for keyword in asset_keywords:
            if keyword in name_upper:
                found.append(keyword)
        
        return ", ".join(found) if found else "-"

    def _format_params(self, params: dict) -> str:
        """Format test parameters for display."""
        if not params:
            return "-"
        
        parts = []
        for key, value in params.items():
            parts.append(f"{key}={value}")
        return ", ".join(parts)

    def _extract_expected(self, docstring: str) -> str:
        """Extract expected behavior from docstring."""
        if not docstring:
            return "-"
        
        keywords = ["should", "must", "expected", "returns", "gives", "==", ">=", "<="]
        for keyword in keywords:
            if keyword.lower() in docstring.lower():
                return docstring.strip()
        
        return docstring.strip() if len(docstring) < 100 else docstring[:97] + "..."

    def _auto_width_columns(self, ws, max_width: int = 40):
        """Auto-adjust column widths based on content."""
        for column_cells in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)
            
            for cell in column_cells:
                try:
                    cell_value = str(cell.value) if cell.value else ""
                    cell_length = len(cell_value)
                    if cell_length > max_length:
                        max_length = cell_length
                except:
                    pass
            
            adjusted_width = min(max_length + 2, max_width)
            ws.column_dimensions[column_letter].width = max(adjusted_width, 10)
